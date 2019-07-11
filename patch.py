import pandas as pd
import openpyxl
import xlsxwriter
from matplotlib import pyplot as plt
from matplotlib import style
import datetime
#function to find sheetnames and workbook name
#def sheet_book():

#function to add data to excel sheet
def append_df_to_excel(filename, df, sheet_name, startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):

    from openpyxl import load_workbook

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl',index = False)

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
        startcolumn = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index = False ,header = False,merge_cells=False,**to_excel_kwargs)
    #print (writer)

    # save the workbook
    writer.save()

#function to read data from csv file
def read_csv(csv,current_time,colname):
    try:

        df1 = pd.read_csv(csv,names = [colname],usecols=[1], skip_blank_lines=False)
        df2=""#declare var df2
        #adding date column to dataframe
        read_csv.current_stat = df1.tail(1)

        #adding pre-defined time to dataframe
        #morning report
        if current_time == 10 or current_time == 11:
            num = 9
            df2 =df1.tail(num).copy()
            df2['time'] = ['17:00','19:00','21:00','23:00','01:00','03:00','05:00','07:00','09:00']
            df2['date'] = pd.to_datetime('today').strftime("%m/%d/%Y")
            df2 = pd.DataFrame(df2, columns = ['date','time',colname])
        #evening report
        elif current_time == 15 or current_time == 16:
            num = 3
            df2 =df1.tail(num).copy()
            df2['time'] = ['11:00','13:00','15:00']
            df2['date'] = pd.to_datetime('today').strftime("%m/%d/%Y")
            df2 = pd.DataFrame(df2, columns = ['date','time',colname])
        else :
            exit()
        #df2['date'] = pd.to_datetime('today').strftime("%m/%d/%Y")
        #df2 = pd.DataFrame(df2, columns = ['date','time','rate'])
        print (df2)
        return df2
    except:
        print("check your csv files please!")

#function to find mean
def find_mean(filename,sheet_name,csv,colname):
    #colname1 = "STKAvgTPS"
    #colname2 = "PeakTPS"
    df = pd.read_excel(filename,sheet_name)
    col_mean = df[colname].mean()
    find_mean.baseline = str(col_mean)
    #col_mean2 = df[colname2].mean()
    print ("avg for this is - " + str(col_mean))

#function to checkout book and sheetnames
def assign():
    sheetname1 = ["STK_Push_TPS","STK_Push_Transaction SR"]
    csvname =[]
    l1 = "SM_STK_Push_TPS-" + str(datetime.date.today()) + ".csv"
    csvname.append(l1)
    l2 =  "SM_STK_Push_Transaction_SR-" + str(datetime.date.today()) + ".csv"
    csvname.append(l2)
    column =["STKAvgTPS","PeakTPS"]
    csv_column=["STKAvgTPS","successRate"]
    #csvname = ["SM_STK_Push_TPS-2019-06-21.csv","SM_STK_Push_Transaction_SR-2019-06-21.csv"]
    for i, j, k, l in zip(sheetname1, csvname, column, csv_column):
        sheet_name = i
        csv = j
        filename = "STK Push NEW.xlsx"
        colname = k
        csvcol = l
        current_time = pd.datetime.now().hour
        df = read_csv(csv,current_time,colname)

        append_df_to_excel(filename , df, sheet_name)
        find_mean(filename,sheet_name,csv,colname)
        draw(filename,sheet_name,csv,colname,csvcol)

        print("*********SUMMARY***********")
        print(j)
        print ("current stat = " + read_csv.current_stat)
        print ("Baseline = " + find_mean.baseline)
def draw(filename,sheet_name,csv, colname, csvcol):
    #draw line graph based on average and current stats

    df = pd.read_excel(filename,sheet_name)
    col_mean = df[colname].mean()


    df5 = pd.read_csv(csv, skip_blank_lines=False)
    df6 =df5.tail()
    #print(df6)
    ax = plt.gca()
    df6.plot(kind = 'line', x= '_time', y= csvcol, ax=ax)
    plt.axhline(col_mean, color='r', linestyle='--')
    name = str(pd.datetime.now()) + 'athman.png'
    #plt.savefig('athman.png')
    plt.savefig(name)
    #plt.show()
    #graph 2
    '''
    #draw line graph based on average and current stats
    df5 = pd.read_csv(csv, skip_blank_lines=False)
    df6 =df5.tail()
    print(df6)
    df6.plot(kind = 'bar', x= '_time', y='STKAvgTPS')
    plt.axhline(col_mean2, color='r', linestyle='--')
    #plt.show()
    '''
assign()
#draw()
